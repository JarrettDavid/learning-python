
# This code is designed to open an Excel spreadsheet, read the data and wrap it in
# HTML

#import 'os'
import os

#import sys
import sys

#import load_workbook module from openpyxl
from openpyxl import load_workbook


#Retrieve current working directory ('cwd')
cwd = os.getcwd()
cwd

#Change directory (not used in this code)
# os.chdir("/path/to/your/folder")

#List all files and directories in current directory (not used in this code)
#os.listdir('.')

#Load the workbook into the variable workbook
workbook = load_workbook('COE-Internal-Faculty-Awards-Recipients-List-08-14-21-jld.xlsx')

#Get the sheet names from the document
print(workbook.sheetnames)

#Get a sheet by name
sheet = workbook['Table 1']

# Converts the department accronym into the full name
def DepartmentFind(CODE):
    dept = {"AERO": "Aerospace","BAEN": "Biological &amp; Agricultural","BMEN": "Biomedical",
            "CHEN": "Chemical","CVEN": "Civil &amp; Environmental","CSCE": "Computer Science",
            "ECEN": "Electrical","ISEN": "Industrial & Systems","MEEN": "Mechanical",
            "MSEN": "Materials Science","MTDE": "Multidisciplinary",
            "NUEN": "Nuclear","OCEN": "Ocean","PETE": "Petroleum"}
    if CODE == "ETID" or CODE == "EASA":
        dept_name = CODE
    else:
        dept_name = dept[CODE]
    return dept_name

# Generates the HTML body for each award Category
def FormatSheet(sheet):
    year_range = ""
    complete_entry = ""
    file_name = "College-Of-Engineering-Award-Recipients.html"
    #section_start = "<section aria-label=“ Recipients” class=” section-wrap--None” role=“contentinfo”><div class=“link-collection-list”><div class=“link-collection-list__item”><div class=“link-collection link-collection--two-column”><h2 class=“link-collection__headline”> Recipients</h2><ul>"
    section_start = ""
    award_entry = ""
    section_end = "</ul></div></div></div></section>"
    for i in range(2, 208):
        cell = 'A' + str(i)
        cell_check = 'D' + str(i)
        if sheet[cell_check].value == None:
            print(sheet[cell].value)
            section = sheet[cell].value
            year_range = section.replace(" College of Engineering Internal Faculty Award Recipients", '')
        else:
            name = 'B' + str(i)
            dept = 'C' + str(i)
            cate = 'D' + str(i)
            pcat = 'D' + str(i-1)
            a_typ = 'E' + str(i)
            ap_typ = 'E' + str(i-1)
            r_name = sheet[name].value
            d_name = sheet[dept].value
            a_cate = sheet[cate].value
            p_cate = sheet[pcat].value
            a_type = sheet[a_typ].value
            ap_type = sheet[ap_typ].value
            year = str(sheet[cell].value)
            
            if p_cate == None:
                if a_type.find(a_cate) == -1 or a_type.find("TEES Research Impact Award:") != -1:
                    award_entry = award_entry + "<li><span><strong>Recipient Name:</strong> "+r_name+"</span><br />"
                    award_entry = award_entry + "<span><strong>Department: </strong> "+DepartmentFind(d_name)+"</span><br />" 
                    award_entry = award_entry + "<span> <strong>Award Name:</strong> "+a_type+"</span><br />"
                    award_entry = award_entry + "<span> <strong>Award Year:</strong> "+year+"</span></li>\n"
                else:
                    award_entry = award_entry + "<li><span><strong>Recipient Name:</strong> "+r_name+"</span><br />"
                    award_entry = award_entry + "<span><strong>Department: </strong> "+DepartmentFind(d_name)+"</span><br />" 
                    award_entry = award_entry + "<span> <strong>Award Year:</strong> "+year+"</span></li>\n"
                if complete_entry != "":
                    print('\n\n\n Write to file main Primary\n',file_name)
                    print('\n\n\n Complete entry Primary\n',complete_entry)
                    if os.path.exists(file_name) == True:
                        sourceFile = open(file_name, 'a')
                    else:
                        sourceFile = open(file_name, 'x')
                    print(complete_entry, file = sourceFile)
                    sourceFile.close()
                    complete_entry = ""
                    award_entry = ""
                new_category = "<h2>Accordion Label: "+year_range.strip()+' '+a_cate.strip()+" Award Recipients</h2>"
                section_start = new_category+"\n<section aria-label=“"+a_type.strip()+" Recipients” class=” section-wrap--None” role=“contentinfo”><div class=“link-collection-list”><div class=“link-collection-list__item”><div class=“link-collection link-collection--two-column”><h2 class=“link-collection__headline”>"+a_cate.strip()+" Recipients</h2><ul>"
            else:
                if p_cate != a_cate:
                    award_entry = section_start +'\n'+ award_entry +'\n'+ section_end + "<br/><h2> End of section, do not paste this into the WYSIWYG</h2><br/>"
                    complete_entry = award_entry;
                    if complete_entry != "":
                        print('\n\n\n Write to file main Secondary\n',file_name)
                        print('\n\n\n Complete entry Secondary\n',complete_entry)
                        if os.path.exists(file_name) == True:
                            sourceFile = open(file_name, 'a')
                        else:
                            sourceFile = open(file_name, 'x')
                        print(complete_entry, file = sourceFile)
                        sourceFile.close()
                        complete_entry = ""
                        award_entry = ""
                    new_category = "<h2>Accordion Label: "+year_range.strip()+' '+a_cate.strip()+" Award Recipients</h2>"
                    print("Writing to file Two: ", file_name)
                    complete_entry = ""
                    award_entry = ""
                    section_start = new_category+"\n<section aria-label=“"+a_type.strip()+" Recipients” class=” section-wrap--None” role=“contentinfo”><div class=“link-collection-list”><div class=“link-collection-list__item”><div class=“link-collection link-collection--two-column”><h2 class=“link-collection__headline”>"+a_type.strip()+" Recipients</h2><ul>"
                else:
                    if a_type.find(a_cate) == -1 or a_type.find("TEES Research Impact Award:") != -1:
                        award_entry = award_entry + "<li><span><strong>Recipient Name:</strong> "+r_name+"</span><br />"
                        award_entry = award_entry + "<span><strong>Department: </strong> "+DepartmentFind(d_name)+"</span><br />" 
                        award_entry = award_entry + "<span> <strong>Award Name:</strong> "+a_type+"</span><br />"
                        award_entry = award_entry + "<span> <strong>Award Year:</strong> "+year+"</span></li>\n"
                    else:
                        award_entry = award_entry + "<li><span><strong>Recipient Name:</strong> "+r_name+"</span><br />"
                        award_entry = award_entry + "<span><strong>Department: </strong> "+DepartmentFind(d_name)+"</span><br />" 
                        award_entry = award_entry + "<span> <strong>Award Year:</strong> "+year+"</span></li>\n"
                    if a_type != ap_type and a_type.find("TEES Research Impact Award:") == -1:
                        award_entry = section_start +'\n'+ award_entry +'\n'+ section_end + "<br/><h2> End of section, do not paste this into the WYSIWYG</h2><br/>"
                        complete_entry = award_entry;
                        if complete_entry != "":
                            print('\n\n\n Write to file main Sub-Cate\n',file_name)
                            print('\n\n\n Complete entry Sub-Cate\n',complete_entry)
                            sourceFile = open(file_name, 'a')
                            print(complete_entry, file = sourceFile)
                            sourceFile.close()
                            complete_entry = ""
                            award_entry = ""
                        new_category_a_type = "<h2>Accordion Label: "+year_range.strip()+' '+a_type.strip()+" Award Type Recipients</h2>"
                        print("Writing to new award type: ", file_name)
                        complete_entry = ""
                        award_entry = ""
                        section_start = new_category_a_type+"\n<section aria-label=“"+a_type.strip()+" Recipients” class=” section-wrap--None” role=“contentinfo”><div class=“link-collection-list”><div class=“link-collection-list__item”><div class=“link-collection link-collection--two-column”><h2 class=“link-collection__headline”>"+a_type.strip()+" Recipients</h2><ul>"
   
    if i == 207:
        award_entry = section_start +'\n'+ award_entry +'\n'+ section_end + "<br/><h2> End of section, do not paste this into the WYSIWYG</h2><br/>"
        complete_entry = award_entry;
        if complete_entry != "":
            print('\n\n\n Write to file main Final\n',file_name)
            print('\n\n\n Complete entry Final\n',complete_entry)
            if os.path.exists(file_name) == True:
                sourceFile = open(file_name, 'a')
            else:
                sourceFile = open(file_name, 'x')
            print(complete_entry, file = sourceFile)
            sourceFile.close()
            complete_entry = ""
            award_entry = ""

FormatSheet(sheet)